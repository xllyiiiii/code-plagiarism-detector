import io
import json
import os
import zipfile
from datetime import datetime

from flask import (Blueprint, render_template, redirect, url_for, flash,
                   request, jsonify, send_file, current_app)
from flask_login import login_required, current_user
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app import db
from app.models import (Assignment, Submission, SimilarityResult,
                         PlagiarismGroup, PlagiarismGroupMember,
                         AnalysisTask, AuditLog)
from app.utils.decorators import role_required
from app.plagiarism.ast_parser import extract_features
from app.plagiarism.similarity import compute_similarity, batch_compare
from app.plagiarism.code_analyzer import analyze_code

plagiarism_bp = Blueprint('plagiarism', __name__)


@plagiarism_bp.route('/')
@login_required
def index():
    """查重任务列表（教师看自己课程的查重记录）。"""
    if current_user.role == 'teacher':
        from app.models import Course
        course_ids = [c.id for c in Course.query.filter_by(
            teacher_id=current_user.id).all()]
        assignments = Assignment.query.filter(
            Assignment.course_id.in_(course_ids)
        ).order_by(Assignment.created_at.desc()).all() if course_ids else []
    elif current_user.role == 'admin':
        assignments = Assignment.query.order_by(
            Assignment.created_at.desc()).all()
    else:
        from app.models import Course
        enrolled_ids = [e.course_id for e in current_user.enrollments.all()]
        assignments = Assignment.query.filter(
            Assignment.course_id.in_(enrolled_ids)
        ).order_by(Assignment.created_at.desc()).all() if enrolled_ids else []

    # Attach analysis task status
    for a in assignments:
        a._latest_task = a.analysis_tasks.order_by(
            AnalysisTask.created_at.desc()).first()

    return render_template('plagiarism/index.html', assignments=assignments)


@plagiarism_bp.route('/run/<int:assignment_id>', methods=['POST'])
@login_required
@role_required('teacher', 'admin')
def run_analysis(assignment_id):
    """启动查重分析（同步执行，适合小规模）。"""
    assignment = Assignment.query.get_or_404(assignment_id)
    submissions = assignment.submissions.order_by(
        Submission.submitted_at.asc()).all()

    if len(submissions) < 2:
        flash('至少需要 2 份提交才能进行查重。')
        return redirect(url_for('plagiarism.index'))

    # Create task record
    task = AnalysisTask(
        assignment_id=assignment.id,
        status='running',
        total_pairs=len(submissions) * (len(submissions) - 1) // 2,
        started_at=datetime.utcnow()
    )
    db.session.add(task)
    db.session.commit()

    try:
        # 1. Parse all submissions
        parsed = []
        for sub in submissions:
            try:
                features = extract_features(sub.file_path, sub.language or 'python')
                sub.ast_data = features
                parsed.append((sub.id, features))
            except Exception as e:
                print(f'[WARN] Failed to parse submission {sub.id}: {e}')
                continue

        if len(parsed) < 2:
            task.status = 'failed'
            task.error_message = '可解析的提交不足 2 份。'
            db.session.commit()
            flash('解析失败：有效的代码文件不足 2 份。')
            return redirect(url_for('plagiarism.index'))

        # 2. Delete old results
        SimilarityResult.query.filter_by(assignment_id=assignment.id).delete()
        PlagiarismGroupMember.query.filter(
            PlagiarismGroupMember.group_id.in_(
                db.session.query(PlagiarismGroup.id).filter_by(
                    assignment_id=assignment.id)
            )
        ).delete(synchronize_session='fetch')
        PlagiarismGroup.query.filter_by(assignment_id=assignment.id).delete()

        # 3. Batch compare
        threshold = assignment.similarity_threshold or 0.70
        results = batch_compare(parsed, threshold=threshold)

        # 4. Store results
        stored_count = 0
        suspicious_subs = set()
        for id_a, id_b, scores in results:
            sim = SimilarityResult(
                assignment_id=assignment.id,
                submission_a_id=id_a,
                submission_b_id=id_b,
                jaccard_similarity=scores['jaccard'],
                tree_edit_similarity=scores['tree_edit'],
                ngram_similarity=scores['ngram'],
                final_similarity=scores['final'],
            )
            db.session.add(sim)
            suspicious_subs.add(id_a)
            suspicious_subs.add(id_b)
            stored_count += 1

        # 5. Group suspicious submissions
        if suspicious_subs:
            group = PlagiarismGroup(
                assignment_id=assignment.id,
                label=f'疑似抄袭组 - {assignment.title[:20]}'
            )
            db.session.add(group)
            db.session.flush()

            # Mark earliest submission as potential original
            earliest_sub = min(
                suspicious_subs,
                key=lambda sid: next(s.submitted_at for s in submissions if s.id == sid)
            )
            for sub_id in suspicious_subs:
                member = PlagiarismGroupMember(
                    group_id=group.id,
                    submission_id=sub_id,
                    is_original=(sub_id == earliest_sub)
                )
                db.session.add(member)

        # 6. Update task
        task.status = 'completed'
        task.processed_pairs = len(parsed) * (len(parsed) - 1) // 2
        task.completed_at = datetime.utcnow()
        db.session.commit()

        flash(f'查重完成！共比对 {task.total_pairs} 对，发现 {stored_count} 对疑似相似。')

    except Exception as e:
        db.session.rollback()
        task.status = 'failed'
        task.error_message = str(e)
        db.session.commit()
        flash(f'查重执行出错：{e}')

    return redirect(url_for('plagiarism.report', assignment_id=assignment.id))


@plagiarism_bp.route('/report/<int:assignment_id>')
@login_required
def report(assignment_id):
    """查看某次作业的查重报告。"""
    assignment = Assignment.query.get_or_404(assignment_id)

    # Get similarity results
    sim_results = assignment.similarity_results.order_by(
        SimilarityResult.final_similarity.desc()).all()

    # Get plagiarism groups
    groups = assignment.plagiarism_groups.all()

    # Get latest analysis task
    task = assignment.analysis_tasks.order_by(
        AnalysisTask.created_at.desc()).first()

    # Build data for heatmap
    submissions = assignment.submissions.order_by(
        Submission.submitted_at.asc()).all()
    labels = [s.student.username for s in submissions]
    heatmap_data = []
    sim_map = {}
    for r in sim_results:
        idx_a = next(i for i, s in enumerate(submissions) if s.id == r.submission_a_id)
        idx_b = next(i for i, s in enumerate(submissions) if s.id == r.submission_b_id)
        heatmap_data.append([idx_a, idx_b, round(r.final_similarity, 2)])
        heatmap_data.append([idx_b, idx_a, round(r.final_similarity, 2)])
        sim_map[(r.submission_a_id, r.submission_b_id)] = r
        sim_map[(r.submission_b_id, r.submission_a_id)] = r

    # Compute dashboard statistics
    if sim_results:
        avg_sim = sum(r.final_similarity for r in sim_results) / len(sim_results)
    else:
        avg_sim = 0.0
    high_risk_count = sum(1 for r in sim_results if r.final_similarity >= 0.85)

    # Build histogram data: 5 bins [0-0.2, 0.2-0.4, 0.4-0.6, 0.6-0.8, 0.8-1.0]
    bins = ['0-20%', '20-40%', '40-60%', '60-80%', '80-100%']
    counts = [0, 0, 0, 0, 0]
    for r in sim_results:
        s = r.final_similarity
        if s < 0.2:
            counts[0] += 1
        elif s < 0.4:
            counts[1] += 1
        elif s < 0.6:
            counts[2] += 1
        elif s < 0.8:
            counts[3] += 1
        else:
            counts[4] += 1

    return render_template('plagiarism/report.html',
                           assignment=assignment,
                           sim_results=sim_results,
                           groups=groups,
                           task=task,
                           submissions=submissions,
                           labels=json.dumps(labels),
                           heatmap_data=json.dumps(heatmap_data),
                           sim_map=sim_map,
                           avg_sim=avg_sim,
                           high_risk_count=high_risk_count,
                           histogram_bins=json.dumps(bins),
                           histogram_counts=json.dumps(counts))


@plagiarism_bp.route('/compare/<int:result_id>')
@login_required
def compare_code(result_id):
    """双栏代码对比视图 —— 含逐行差异标注。"""
    sim = SimilarityResult.query.get_or_404(result_id)

    def read_code(submission):
        try:
            with open(submission.file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
        except Exception:
            return '[无法读取文件]'

    code_a = read_code(sim.submission_a)
    code_b = read_code(sim.submission_b)
    lang = (sim.submission_a.language or 'python')

    # Map language names to highlight.js classes
    lang_map = {'python': 'python', 'java': 'java', 'c': 'c', 'cpp': 'cpp',
                'javascript': 'javascript', 'js': 'javascript'}
    hl_lang = lang_map.get(lang, 'plaintext')

    # Compute line-level diff
    import difflib
    lines_a_raw = code_a.splitlines()
    lines_b_raw = code_b.splitlines()
    sm = difflib.SequenceMatcher(None, lines_a_raw, lines_b_raw)
    opcodes = sm.get_opcodes()

    # Build highlighted line lists
    lines_a = []
    lines_b = []
    for tag, i1, i2, j1, j2 in opcodes:
        if tag == 'equal':
            for k in range(i1, i2):
                lines_a.append({'num': k + 1, 'text': lines_a_raw[k], 'css_class': '', 'sign': ' '})
            for k in range(j1, j2):
                lines_b.append({'num': k + 1, 'text': lines_b_raw[k], 'css_class': '', 'sign': ' '})
        elif tag == 'replace':
            for k in range(i1, i2):
                lines_a.append({'num': k + 1, 'text': lines_a_raw[k], 'css_class': 'diff-replace', 'sign': '~'})
            for k in range(j1, j2):
                lines_b.append({'num': k + 1, 'text': lines_b_raw[k], 'css_class': 'diff-replace', 'sign': '~'})
        elif tag == 'delete':
            for k in range(i1, i2):
                lines_a.append({'num': k + 1, 'text': lines_a_raw[k], 'css_class': 'diff-delete', 'sign': '-'})
            for k in range(j1, j2):
                pass  # no corresponding lines in B (but we keep empty placeholders for alignment)
        elif tag == 'insert':
            for k in range(i1, i2):
                pass
            for k in range(j1, j2):
                lines_b.append({'num': k + 1, 'text': lines_b_raw[k], 'css_class': 'diff-insert', 'sign': '+'})

    # Log audit
    log = AuditLog(
        user_id=current_user.id,
        action='view_code_compare',
        target_type='similarity_result',
        target_id=result_id,
        ip_address=request.remote_addr,
        detail=f'compare {sim.submission_a_id} vs {sim.submission_b_id}'
    )
    db.session.add(log)
    db.session.commit()

    return render_template('plagiarism/compare.html', sim=sim,
                           lines_a=lines_a, lines_b=lines_b,
                           lang=hl_lang)


@plagiarism_bp.route('/review/<int:result_id>', methods=['POST'])
@login_required
@role_required('teacher', 'admin')
def review(result_id):
    """教师人工复核标注。"""
    sim = SimilarityResult.query.get_or_404(result_id)
    action = request.form.get('action')  # confirm / clear / false_positive

    if action == 'confirm':
        sim.is_plagiarism = True
    elif action == 'clear':
        sim.is_plagiarism = False
    else:
        flash('无效的复核操作。')
        return redirect(request.referrer or url_for('plagiarism.index'))

    sim.reviewed_by = current_user.id
    sim.reviewed_at = datetime.utcnow()
    db.session.commit()

    flash('复核结果已保存。')
    return redirect(url_for('plagiarism.report',
                            assignment_id=sim.assignment_id))


@plagiarism_bp.route('/api/graph/<int:assignment_id>')
@login_required
def api_graph_data(assignment_id):
    """API: 返回图谱数据（供 ECharts 渲染）。"""
    assignment = Assignment.query.get_or_404(assignment_id)
    submissions = assignment.submissions.all()
    results = assignment.similarity_results.filter(
        SimilarityResult.final_similarity >= (assignment.similarity_threshold or 0.70)
    ).all()

    nodes = []
    for s in submissions:
        nodes.append({
            'id': s.id,
            'name': s.student.display_name or s.student.username,
            'symbolSize': 30,
        })

    links = []
    for r in results:
        links.append({
            'source': r.submission_a_id,
            'target': r.submission_b_id,
            'value': round(r.final_similarity, 2),
        })

    return jsonify({'nodes': nodes, 'links': links})


# ================================================================
# 学生端：学习辅助
# ================================================================

@plagiarism_bp.route('/my-report')
@login_required
@role_required('student')
def my_report():
    """学生个人学习报告 —— 历史相似度趋势 + 代码质量得分。"""
    submissions = current_user.submissions.order_by(
        Submission.submitted_at.desc()).all()

    # Analyze each submission
    analyzed = []
    for sub in submissions:
        try:
            report = analyze_code(sub.file_path, sub.language or 'python')
        except Exception:
            report = {'score': None, 'grade': 'N/A', 'issues': [], 'suggestions': []}
        analyzed.append({
            'submission': sub,
            'report': report,
        })

    # Get similarity results involving this student
    from sqlalchemy import or_
    sim_results = SimilarityResult.query.filter(
        or_(
            SimilarityResult.submission_a_id.in_([s.id for s in submissions]),
            SimilarityResult.submission_b_id.in_([s.id for s in submissions]),
        )
    ).order_by(SimilarityResult.final_similarity.desc()).all()

    # Trend data for chart
    dates = []
    scores = []
    sim_trends = []
    for sub in sorted(submissions, key=lambda s: s.submitted_at):
        dates.append(sub.submitted_at.strftime('%m-%d'))
        # Get quality score
        try:
            r = analyze_code(sub.file_path, sub.language or 'python')
            scores.append(r['score'])
        except Exception:
            scores.append(0)
        # Get max similarity for this submission
        max_sim = 0
        for sr in sim_results:
            if sr.submission_a_id == sub.id or sr.submission_b_id == sub.id:
                max_sim = max(max_sim, sr.final_similarity)
        sim_trends.append(round(max_sim * 100, 1))

    return render_template('plagiarism/student_report.html',
                           analyzed=analyzed,
                           submissions=submissions,
                           sim_results=sim_results,
                           dates=json.dumps(dates),
                           scores=json.dumps(scores),
                           sim_trends=json.dumps(sim_trends))


@plagiarism_bp.route('/code-review/<int:submission_id>')
@login_required
def code_review(submission_id):
    """代码质量详细审查报告。"""
    sub = Submission.query.get_or_404(submission_id)

    # Permission: student can only see own, teacher/admin can see all
    if current_user.role == 'student' and sub.student_id != current_user.id:
        flash('无权查看此报告。')
        return redirect(url_for('plagiarism.my_report'))

    try:
        report = analyze_code(sub.file_path, sub.language or 'python')
    except Exception as e:
        flash(f'代码分析失败：{e}')
        return redirect(request.referrer or url_for('plagiarism.index'))

    # Get similar submissions for context
    sim_pairs = SimilarityResult.query.filter(
        (SimilarityResult.submission_a_id == submission_id) |
        (SimilarityResult.submission_b_id == submission_id)
    ).order_by(SimilarityResult.final_similarity.desc()).all()

    # Generate pair-specific refactoring suggestions
    refactor_suggestions = []
    for sr in sim_pairs:
        other_sub = sr.submission_b if sr.submission_a_id == submission_id else sr.submission_a
        suggestions = _generate_pair_suggestions(sr, sub, other_sub)
        if suggestions:
            refactor_suggestions.append({
                'similarity_result': sr,
                'other_student': other_sub.student.display_name or other_sub.student.username,
                'other_file': other_sub.file_name,
                'suggestions': suggestions,
            })

    try:
        with open(sub.file_path, 'r', encoding='utf-8', errors='ignore') as f:
            source_code = f.read()
    except Exception:
        source_code = '[无法读取]'

    return render_template('plagiarism/code_review.html',
                           submission=sub,
                           report=report,
                           sim_pairs=sim_pairs,
                           source_code=source_code,
                           refactor_suggestions=refactor_suggestions)


def _generate_pair_suggestions(sim_result, sub_a, sub_b):
    """Generate specific refactoring suggestions based on similarity pattern."""
    suggestions = []
    final_sim = sim_result.final_similarity
    jaccard = sim_result.jaccard_similarity or 0
    tree_edit = sim_result.tree_edit_similarity or 0
    ngram = sim_result.ngram_similarity or 0

    if final_sim < 0.6:
        return suggestions

    # High Jaccard = similar structure with different names (likely renaming)
    if jaccard > 0.7:
        suggestions.append({
            'title': '结构高度雷同',
            'detail': 'AST 子树相似度很高，说明代码结构几乎一致。建议重新设计算法逻辑，采用不同的数据结构或算法策略来解决问题。',
            'resource': '《算法导论》- 同一问题通常有多种等价的算法实现方式'
        })

    # High tree edit = almost identical structure (minor modifications)
    if tree_edit > 0.7:
        suggestions.append({
            'title': '逻辑结构相同',
            'detail': '树编辑距离分析表明代码逻辑结构与参照代码高度一致。建议提取不同的函数划分、使用不同的循环/递归策略来表达相同逻辑。',
            'resource': '《重构》第6章 - 提炼函数与算法替换'
        })

    # High ngram but lower tree edit = text-level copying with modifications
    if ngram > 0.7 and tree_edit < 0.6:
        suggestions.append({
            'title': '疑似修改抄袭',
            'detail': 'Token 序列高度相似但结构有差异，可能在原代码基础上进行了局部修改。建议从根本上重新设计实现方案，而非在他人代码上修修补补。',
            'resource': '《代码整洁之道》第1章 - 写出有意义的代码'
        })

    # Very high similarity overall
    if final_sim >= 0.85:
        suggestions.append({
            'title': '严重相似警告',
            'detail': f'与 {sub_b.student.display_name or sub_b.student.username} 的代码高度相似({final_sim*100:.0f}%)。'
                       '建议：1) 重新独立完成该作业；2) 与老师沟通说明情况；3) 学习下方推荐的资料后重写代码。',
            'resource': '请独立完成作业，遇到困难可在课程论坛或答疑时间寻求帮助'
        })

    # Moderate similarity
    if 0.6 <= final_sim < 0.85:
        suggestions.append({
            'title': '适度重构建议',
            'detail': '代码存在一定相似度但仍有改进空间。建议：封装不同的工具类/函数、采用不同的变量命名体系、调整代码组织结构。',
            'resource': '《程序员修炼之道》- DRY 原则与正交性'
        })

    return suggestions


# ================================================================
# 报告导出
# ================================================================

@plagiarism_bp.route('/export/<int:assignment_id>')
@login_required
@role_required('teacher', 'admin')
def export_excel(assignment_id):
    """导出查重结果为 Excel 文件。"""
    assignment = Assignment.query.get_or_404(assignment_id)
    sim_results = assignment.similarity_results.order_by(
        SimilarityResult.final_similarity.desc()).all()
    submissions = {s.id: s for s in assignment.submissions.all()}

    wb = Workbook()
    ws = wb.active
    ws.title = '查重结果'

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    red_font = Font(color='D32F2F', bold=True)
    orange_font = Font(color='F57C00', bold=True)
    green_font = Font(color='388E3C', bold=True)

    headers = ['序号', '学生 A', '文件 A', '学生 B', '文件 B',
               'Jaccard相似度', '树编辑相似度', 'N-gram相似度', '综合相似度',
               '复核状态']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, r in enumerate(sim_results, 2):
        a = submissions.get(r.submission_a_id)
        b = submissions.get(r.submission_b_id)
        row_data = [
            i - 1,
            a.student.display_name or a.student.username if a else '?',
            a.file_name if a else '?',
            b.student.display_name or b.student.username if b else '?',
            b.file_name if b else '?',
            f'{r.jaccard_similarity * 100:.1f}%' if r.jaccard_similarity else 'N/A',
            f'{r.tree_edit_similarity * 100:.1f}%' if r.tree_edit_similarity else 'N/A',
            f'{r.ngram_similarity * 100:.1f}%' if r.ngram_similarity else 'N/A',
            f'{r.final_similarity * 100:.1f}%',
            '待复核' if r.is_plagiarism is None else '已确认抄袭' if r.is_plagiarism else '人工排除',
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if col == 9:
                sim_val = r.final_similarity
                cell.font = red_font if sim_val >= 0.85 else orange_font if sim_val >= 0.70 else green_font

    # Column widths
    widths = [6, 16, 22, 16, 22, 14, 14, 14, 14, 12]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    # Summary sheet
    ws2 = wb.create_sheet('摘要')
    ws2.merge_cells('A1:B1')
    c = ws2.cell(row=1, column=1, value=f'查重报告：{assignment.title}')
    c.font = Font(bold=True, size=14)
    summary_rows = [
        ('课程', assignment.course.name),
        ('提交数', len(submissions)),
        ('相似度阈值', f'{assignment.similarity_threshold * 100:.0f}%'),
        ('发现相似对', len(sim_results)),
        ('高风险对(≥85%)', sum(1 for r in sim_results if r.final_similarity >= 0.85)),
        ('平均相似度', f'{sum(r.final_similarity for r in sim_results) / max(len(sim_results), 1) * 100:.1f}%' if sim_results else 'N/A'),
    ]
    for i, (label, value) in enumerate(summary_rows, 2):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=str(value))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'查重报告_{assignment.title}_{datetime.utcnow().strftime("%Y%m%d")}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@plagiarism_bp.route('/download/<int:assignment_id>')
@login_required
@role_required('teacher', 'admin')
def download_all(assignment_id):
    """一键打包下载某次作业的全部提交。"""
    assignment = Assignment.query.get_or_404(assignment_id)
    submissions = assignment.submissions.all()

    if not submissions:
        flash('该作业暂无提交文件。')
        return redirect(url_for('plagiarism.report', assignment_id=assignment.id))

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for sub in submissions:
            student_name = sub.student.display_name or sub.student.username
            arcname = f'{student_name}/{sub.file_name}'
            if os.path.exists(sub.file_path):
                zf.write(sub.file_path, arcname)
            else:
                zf.writestr(f'{student_name}/_MISSING_.txt',
                            f'文件缺失: {sub.file_path}')

    buf.seek(0)
    filename = f'{assignment.title}_提交汇总_{datetime.utcnow().strftime("%Y%m%d")}.zip'
    return send_file(buf, mimetype='application/zip',
                     as_attachment=True, download_name=filename)
